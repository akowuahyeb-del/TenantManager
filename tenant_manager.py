"""
Tenant Management Excel Workbook
=================================
Sheets:
  1. DASHBOARD   - live KPIs, expiry alerts, occupancy
  2. TENANT FORM - data-entry form styled like the Student Database mockup
  3. TENANTS     - master register with all tenant details
  4. AGREEMENTS  - lease terms, renewals, payment tracking
  5. ROOMS       - room/unit occupancy lookup
  6. LOOKUP      - drop-down lists (hidden helper sheet)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

OUT = "Tenant_Manager.xlsx"

# ---------------------------------------------------------------
C = {
    "navy":        "1F3864",
    "teal":        "17A589",
    "amber":       "E67E22",
    "red":         "E74C3C",
    "green":       "1E8449",
    "light_teal":  "D1F2EB",
    "light_amber": "FDEBD0",
    "light_red":   "FADBD8",
    "light_green": "D5F5E3",
    "light_blue":  "D6EAF8",
    "white":       "FFFFFF",
    "off_white":   "F8F9FA",
    "mid_grey":    "BFC9CA",
    "dark_grey":   "566573",
    "yellow_inp":  "FFF9C4",
    "label_bg":    "FDEBD5",   # peach, like the Student Database mockup
    "panel_bg":    "FDEBD5",
    "outer_bg":    "F8D7CE",
    "btn_bg":      "EDEDED",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def border(style="thin", color="BFC9CA"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thin = Side(border_style="thin", color="BFC9CA")
    thk = Side(border_style="medium", color="1F3864")
    return Border(left=thin, right=thin, top=thin, bottom=thk)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def alternating(ws, row, col_start, col_end, even):
    bg = C["off_white"] if even else C["white"]
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill(bg)
        c.border = border()
        c.alignment = align()

def add_validation(ws, col_letter, row_start, row_end, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)

# =================================================================
wb = openpyxl.Workbook()

# =================================================================
# LOOKUP (hidden helper)
# =================================================================
lk = wb.active
lk.title = "LOOKUP"

UNIT_TYPES  = ["Apartment", "Office", "Shop", "Warehouse", "Studio", "Penthouse", "Other"]
STATUS_LIST = ["Active", "Pending", "Expired", "Terminated", "Vacant"]
FREQ_LIST   = ["Monthly", "Quarterly", "Semi-Annual", "Annual"]
CURRENCY    = ["GHS", "USD", "GBP", "EUR", "NGN"]
PAYMENT_M   = ["Cash", "Bank Transfer", "Mobile Money", "Cheque", "Standing Order"]
YES_NO      = ["Yes", "No"]
SEARCH_FILTER = ["All", "Active", "Expiring Soon", "Expired", "Vacant"]

lists = {
    "UnitTypes": UNIT_TYPES, "StatusList": STATUS_LIST, "FreqList": FREQ_LIST,
    "Currency": CURRENCY, "PayMethod": PAYMENT_M, "YesNo": YES_NO,
    "SearchFilter": SEARCH_FILTER,
}

lk["A1"].value, lk["B1"].value = "List Name", "Values"
r = 2
ranges = {}
for name, items in lists.items():
    start = r
    for item in items:
        lk.cell(row=r, column=2).value = item
        r += 1
    end = r - 1
    ranges[name] = (start, end)
    lk.cell(row=start, column=1).value = name
lk.sheet_state = "hidden"

def lookup_range(name):
    s, e = ranges[name]
    return f"LOOKUP!$B${s}:$B${e}"

MAX_ROWS = 53  # TENANTS data rows: 4..53

# =================================================================
# TENANT FORM  (redesigned to match the Student Database mockup)
# =================================================================
form = wb.create_sheet("TENANT FORM")
form.sheet_view.showGridLines = False
form.sheet_properties.tabColor = C["teal"]

for col, w in zip("ABCDEFGHIJKL", [2, 16, 24, 3, 16, 24, 3, 16, 24, 2, 16, 2]):
    form.column_dimensions[col].width = w

# Outer page background band
for row in range(1, 46):
    for col in range(1, 13):
        form.cell(row=row, column=col).fill = fill(C["outer_bg"])

# --- Title bar ---
form.merge_cells("A1:L1")
c = form["A1"]
c.value = "Tenant Database"
c.fill = fill(C["dark_grey"])
c.font = font(bold=True, color=C["white"], size=16)
c.alignment = align("center")
form.row_dimensions[1].height = 30

# --- Frame1 panel (peach box like the mockup) ---
FRAME_TOP, FRAME_BOTTOM = 3, 30
form.merge_cells(f"A{FRAME_TOP}:L{FRAME_TOP}")
c = form[f"A{FRAME_TOP}"]
c.value = "Frame1"
c.font = font(bold=True, size=10)
c.fill = fill(C["outer_bg"])

for row in range(FRAME_TOP + 1, FRAME_BOTTOM + 1):
    for col in range(1, 13):
        cell = form.cell(row=row, column=col)
        cell.fill = fill(C["panel_bg"])

def form_label(cell_ref, text):
    cell = form[cell_ref]
    cell.value = text
    cell.font = font(bold=True, color=C["navy"], size=10)
    cell.fill = fill(C["panel_bg"])
    cell.alignment = align("left")

def input_box(cell_ref, merge_to=None, value=None, readonly_note=False):
    cell = form[cell_ref]
    if merge_to:
        form.merge_cells(f"{cell_ref}:{merge_to}")
    cell.fill = fill(C["yellow_inp"] if not readonly_note else C["off_white"])
    cell.border = border(style="medium", color="808080")
    cell.font = font(size=10, color="00008B")
    cell.alignment = align()
    if value is not None:
        cell.value = value

# --- Column 1: identity fields ---
col1 = [
    (5,  "Tenant ID",           "B5",  "C5",  "T-001"),
    (7,  "Full Name / Company", "B7",  "C7",  None),
    (9,  "Phone",               "B9",  "C9",  None),
    (11, "Email",                "B11", "C11", None),
    (13, "ID / Reg No.",        "B13", "C13", None),
    (15, "Unit Number",         "B15", "C15", None),
    (17, "Unit Type",           "B17", "C17", None),
    (19, "Floor / Block",       "B19", "C19", None),
]
for row, label, lcell, icell, val in col1:
    form_label(lcell, label)
    input_box(icell, value=val, readonly_note=(label == "Tenant ID"))

# --- Column 2: lease / financial fields ---
col2 = [
    (5,  "Contract Start",       "E5",  "F5"),
    (7,  "Contract End",         "E7",  "F7"),
    (9,  "Monthly Rent",         "E9",  "F9"),
    (11, "Currency",             "E11", "F11"),
    (13, "Payment Frequency",    "E13", "F13"),
    (15, "Deposit Paid",         "E15", "F15"),
    (17, "Payment Method",       "E17", "F17"),
    (19, "Residential Address",  "E19", "F19"),
]
for row, label, lcell, icell in col2:
    form_label(lcell, label)
    input_box(icell)

# --- Column 3: emergency contact + photo placeholder ---
col3 = [
    (5,  "Emergency Contact", "H5", "I5"),
    (7,  "Relationship",      "H7", "I7"),
    (9,  "Emergency Phone",   "H9", "I9"),
]
for row, label, lcell, icell in col3:
    form_label(lcell, label)
    input_box(icell)

# Photo / document placeholder box (like the picture box in the mockup)
form.merge_cells("H12:K19")
photo = form["H12"]
photo.fill = fill(C["off_white"])
photo.border = border(style="medium", color="808080")
photo.alignment = align("center", wrap=True)
photo.value = "Tenant ID photo / lease scan"
photo.font = font(size=9, color=C["dark_grey"], italic=True)

form.merge_cells("H21:K22")
upload_btn = form["H21"]
upload_btn.value = "Upload Document"
upload_btn.fill = fill(C["btn_bg"])
upload_btn.font = font(bold=True, size=10)
upload_btn.alignment = align("center")
upload_btn.border = border(style="thin", color="808080")

# --- Dropdown validations on the form ---
add_validation(form, "C17", 17, 17, lookup_range("UnitTypes"))
add_validation(form, "F11", 11, 11, lookup_range("Currency"))
add_validation(form, "F13", 13, 13, lookup_range("FreqList"))
add_validation(form, "F17", 17, 17, lookup_range("PayMethod"))

for row_addr in ["C5", "F5", "F7"]:
    pass
form["E5"].number_format = "DD-MMM-YYYY"
form["E7"].number_format = "DD-MMM-YYYY"
form["F5"].number_format = "DD-MMM-YYYY"
form["F7"].number_format = "DD-MMM-YYYY"
form["F9"].number_format = "#,##0.00"
form["F15"].number_format = "#,##0.00"

# --- Search row ---
SEARCH_ROW = 25
form_label(f"B{SEARCH_ROW}", "Search")
form.merge_cells(f"C{SEARCH_ROW}:F{SEARCH_ROW}")
search_box = form[f"C{SEARCH_ROW}"]
search_box.fill = fill(C["yellow_inp"])
search_box.border = border(style="medium", color="808080")
search_box.alignment = align()

form.merge_cells(f"G{SEARCH_ROW}:H{SEARCH_ROW}")
filt = form[f"G{SEARCH_ROW}"]
filt.value = "All"
filt.fill = fill(C["white"])
filt.border = border(style="medium", color="808080")
filt.alignment = align("center")
add_validation(form, "G", SEARCH_ROW, SEARCH_ROW, lookup_range("SearchFilter"))

form.merge_cells(f"I{SEARCH_ROW}:J{SEARCH_ROW}")
search_btn = form[f"I{SEARCH_ROW}"]
search_btn.value = "Search"
search_btn.fill = fill(C["btn_bg"])
search_btn.font = font(bold=True, size=10)
search_btn.alignment = align("center")
search_btn.border = border(style="thin", color="808080")

form.row_dimensions[SEARCH_ROW].height = 22

# --- Database preview grid (mirrors the yellow-header grid in the mockup) ---
GRID_TITLE_ROW = 32
form.merge_cells(f"A{GRID_TITLE_ROW}:L{GRID_TITLE_ROW}")
gtitle = form[f"A{GRID_TITLE_ROW}"]
gtitle.value = "Database"
gtitle.fill = fill(C["outer_bg"])
gtitle.font = font(bold=True, size=11)
gtitle.alignment = align("left")

GRID_HEADER_ROW = 33
GRID_HEADERS = [
    ("A", "Tenant ID", 11), ("B", "Name", 20), ("C", "Unit", 9),
    ("D", "Rent", 10), ("E", "Currency", 9), ("F", "Lease End", 12),
    ("G", "Status", 12), ("H", "Phone", 13), ("I", "Emergency Contact", 16),
    ("J", "Emergency Phone", 14), ("K", "Address", 18), ("L", "Payment Method", 14),
]
for col_letter, header, width in GRID_HEADERS:
    c = form[f"{col_letter}{GRID_HEADER_ROW}"]
    c.value = header
    c.fill = fill("FFFF99")
    c.font = font(bold=True, size=9)
    c.border = border()
    c.alignment = align("center")
    form.column_dimensions[col_letter].width = max(form.column_dimensions[col_letter].width or 0, width)

GRID_ROWS = 6
for i in range(GRID_ROWS):
    row = GRID_HEADER_ROW + 1 + i
    src = 4 + i  # TENANTS data starts row 4
    mapping = {
        "A": f"=IFERROR(INDEX(TENANTS!A$4:A${MAX_ROWS},{i+1}),\"\")",
        "B": f"=IFERROR(INDEX(TENANTS!B$4:B${MAX_ROWS},{i+1}),\"\")",
        "C": f"=IFERROR(INDEX(TENANTS!F$4:F${MAX_ROWS},{i+1}),\"\")",
        "D": f"=IFERROR(INDEX(TENANTS!L$4:L${MAX_ROWS},{i+1}),\"\")",
        "E": f"=IFERROR(INDEX(TENANTS!M$4:M${MAX_ROWS},{i+1}),\"\")",
        "F": f"=IFERROR(INDEX(TENANTS!J$4:J${MAX_ROWS},{i+1}),\"\")",
        "G": f"=IFERROR(INDEX(TENANTS!R$4:R${MAX_ROWS},{i+1}),\"\")",
        "H": f"=IFERROR(INDEX(TENANTS!C$4:C${MAX_ROWS},{i+1}),\"\")",
        "I": f"=IFERROR(INDEX(TENANTS!S$4:S${MAX_ROWS},{i+1}),\"\")",
        "J": f"=IFERROR(INDEX(TENANTS!U$4:U${MAX_ROWS},{i+1}),\"\")",
        "K": f"=IFERROR(INDEX(TENANTS!V$4:V${MAX_ROWS},{i+1}),\"\")",
        "L": f"=IFERROR(INDEX(TENANTS!P$4:P${MAX_ROWS},{i+1}),\"\")",
    }
    for col_letter, formula in mapping.items():
        c = form[f"{col_letter}{row}"]
        c.value = formula
        c.font = font(size=9)
        c.border = border()
        c.alignment = align()
        c.fill = fill(C["off_white"] if i % 2 else C["white"])
    form[f"D{row}"].number_format = "#,##0.00"
    form[f"F{row}"].number_format = "DD-MMM-YYYY"

# --- Button row (Add / Update / Delete / Reset / Close) ---
BTN_ROW = GRID_HEADER_ROW + GRID_ROWS + 2
btn_defs = [
    ("A", "B", "Add"),
    ("D", "E", "Update"),
    ("G", "H", "Delete"),
    ("I", "J", "Reset"),
    ("K", "L", "Close"),
]
for cs, ce, label in btn_defs:
    form.merge_cells(f"{cs}{BTN_ROW}:{ce}{BTN_ROW}")
    b = form[f"{cs}{BTN_ROW}"]
    b.value = label
    b.fill = fill(C["btn_bg"])
    b.font = font(bold=True, size=11)
    b.alignment = align("center")
    b.border = border(style="thin", color="808080")
form.row_dimensions[BTN_ROW].height = 24

form.merge_cells(f"A{BTN_ROW+2}:L{BTN_ROW+2}")
note = form[f"A{BTN_ROW+2}"]
note.value = "Fill the yellow cells, then use Add to append a new tenant to the TENANTS sheet, or Update / Delete for the selected row."
note.font = font(size=8, italic=True, color=C["dark_grey"])
note.alignment = align("center")

# =================================================================
# TENANTS
# =================================================================
tn = wb.create_sheet("TENANTS")
tn.sheet_view.showGridLines = False

tn.merge_cells("A1:V1")
c = tn["A1"]
c.value = "TENANT REGISTER"
c.fill = fill(C["navy"])
c.font = font(bold=True, color=C["white"], size=14)
c.alignment = align("center")
tn.row_dimensions[1].height = 32

tn.merge_cells("A2:V2")
c = tn["A2"]
c.value = "Yellow cells = data entry - Blue text = formula (do not edit) - Status updates automatically"
c.fill = fill(C["dark_grey"])
c.font = font(color=C["white"], size=9, italic=True)
c.alignment = align("center")
tn.row_dimensions[2].height = 18

T_HEADERS = [
    ("A", "Tenant ID", 11), ("B", "Full Name / Company", 22),
    ("C", "Phone", 14), ("D", "Email", 24),
    ("E", "ID / Reg No.", 14), ("F", "Unit No.", 10),
    ("G", "Unit Type", 12), ("H", "Floor / Block", 12),
    ("I", "Contract Start", 13), ("J", "Contract End", 13),
    ("K", "Term (months)", 13), ("L", "Monthly Rent", 13),
    ("M", "Currency", 10), ("N", "Payment Freq.", 13),
    ("O", "Deposit Paid", 13), ("P", "Payment Method", 14),
    ("Q", "Days to Expiry", 14), ("R", "Status", 12),
    ("S", "Emergency Contact", 20), ("T", "Relationship", 15),
    ("U", "Emergency Phone", 16), ("V", "Residential Address", 25),
]

for col_letter, header, width in T_HEADERS:
    c = tn[f"{col_letter}3"]
    c.value = header
    c.fill = fill(C["navy"])
    c.font = font(bold=True, color=C["white"], size=9)
    c.alignment = align("center")
    c.border = thick_bottom()
    tn.column_dimensions[col_letter].width = width

tn.row_dimensions[3].height = 22
tn.freeze_panes = "A4"

EXAMPLE_ROWS = [
    {"A": "T-001", "B": "Kwame Mensah Enterprises", "C": "0244123456", "D": "kwame@example.com",
     "E": "GH-0012345", "F": "A-101", "G": "Office", "H": "1st / Block A",
     "I": datetime.date(2024, 1, 1), "J": datetime.date(2024, 12, 31), "L": 3500, "M": "GHS",
     "N": "Monthly", "O": 7000, "P": "Bank Transfer",
     "S": "Ama Mensah", "T": "Spouse", "U": "0244999111", "V": "12 Ridge Road, Kumasi"},
    {"A": "T-002", "B": "Simeon Boison", "C": "0598833244", "D": "simeonb@example.com",
     "E": "GH-0022341", "F": "B-204", "G": "Apartment", "H": "2nd / Block B",
     "I": datetime.date(2025, 8, 15), "J": datetime.date(2026, 8, 15), "L": 2800, "M": "GHS",
     "N": "Monthly", "O": 5600, "P": "Mobile Money",
     "S": "Yaw Boison", "T": "Brother", "U": "0209988112", "V": "5 Adum Street, Kumasi"},
    {"A": "T-003", "B": "Benjamin Darko", "C": "0545921808", "D": "dynamicdb2541@gmail.com",
     "E": "GH-0033452", "F": "C-010", "G": "Shop", "H": "Ground / Block C",
     "I": datetime.date(2025, 1, 2), "J": datetime.date(2026, 1, 2), "L": 4100, "M": "GHS",
     "N": "Monthly", "O": 8200, "P": "Cash",
     "S": "Grace Darko", "T": "Sister", "U": "0271234567", "V": "22 Tafo Road, Kumasi"},
]

input_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "L", "M", "N", "O", "P", "S", "T", "U", "V"]

for row in range(4, MAX_ROWS + 1):
    idx = row - 4
    is_example = idx < len(EXAMPLE_ROWS)
    is_even = (row % 2 == 0)

    for col_letter, _, _ in T_HEADERS:
        c = tn[f"{col_letter}{row}"]
        c.font = font(size=9)
        c.border = border()
        c.alignment = align()
        if col_letter in input_cols:
            if is_example and col_letter in EXAMPLE_ROWS[idx]:
                c.value = EXAMPLE_ROWS[idx][col_letter]
            c.fill = fill(C["yellow_inp"]) if is_example else fill(C["off_white"] if is_even else C["white"])
            c.font = font(color="00008B" if is_example else "000000", size=9)

    k = tn[f"K{row}"]
    k.value = f'=IFERROR(DATEDIF(I{row},J{row},"M"),"")'
    k.font = font(color="1A5276", size=9)
    k.alignment = align("center")
    k.fill = fill(C["light_blue"])

    q = tn[f"Q{row}"]
    q.value = f'=IFERROR(IF(J{row}="","",J{row}-TODAY()),"")'
    q.font = font(color="1A5276", size=9)
    q.alignment = align("center")
    q.number_format = "0"
    q.fill = fill(C["light_blue"])

    r_cell = tn[f"R{row}"]
    r_cell.value = (
        f'=IFERROR(IF(B{row}="","Vacant",'
        f'IF(Q{row}<0,"Expired",'
        f'IF(Q{row}<=30,"Expiring Soon",'
        f'IF(Q{row}<=90,"Active - Review","Active")))),"Vacant")'
    )
    r_cell.font = font(bold=True, size=9)
    r_cell.alignment = align("center")
    r_cell.fill = fill(C["light_blue"])

    tn[f"I{row}"].number_format = "DD-MMM-YYYY"
    tn[f"J{row}"].number_format = "DD-MMM-YYYY"
    tn[f"L{row}"].number_format = "#,##0.00"
    tn[f"O{row}"].number_format = "#,##0.00"

red_rule = FormulaRule(formula=['=OR(R4="Expired",R4="Terminated")'], fill=fill(C["light_red"]), font=font(bold=True, color=C["red"]))
amber_rule = FormulaRule(formula=['=R4="Expiring Soon"'], fill=fill(C["light_amber"]), font=font(bold=True, color=C["amber"]))
green_rule = FormulaRule(formula=['=OR(R4="Active",R4="Active - Review")'], fill=fill(C["light_green"]), font=font(bold=True, color=C["green"]))
tn.conditional_formatting.add(f"R4:R{MAX_ROWS}", red_rule)
tn.conditional_formatting.add(f"R4:R{MAX_ROWS}", amber_rule)
tn.conditional_formatting.add(f"R4:R{MAX_ROWS}", green_rule)

tn.conditional_formatting.add(
    f"Q4:Q{MAX_ROWS}",
    ColorScaleRule(start_type="num", start_value=0, start_color="E74C3C",
                   mid_type="num", mid_value=90, mid_color="F39C12",
                   end_type="num", end_value=365, end_color="27AE60"),
)

add_validation(tn, "G", 4, MAX_ROWS, lookup_range("UnitTypes"))
add_validation(tn, "M", 4, MAX_ROWS, lookup_range("Currency"))
add_validation(tn, "N", 4, MAX_ROWS, lookup_range("FreqList"))
add_validation(tn, "P", 4, MAX_ROWS, lookup_range("PayMethod"))

# =================================================================
# ROOMS
# =================================================================
rm = wb.create_sheet("ROOMS")
ROOM_HEADERS = [("A", "Room No", 12), ("B", "Tenant ID", 12), ("C", "Tenant Name", 25),
                ("D", "Monthly Rent", 15), ("E", "Status", 15)]
for col, header, width in ROOM_HEADERS:
    rm[f"{col}1"] = header
    rm.column_dimensions[col].width = width

for i in range(1, 13):
    row = i + 1
    rm[f"A{row}"] = f"RM-{i:02d}"
    rm[f"C{row}"] = f'=IFERROR(INDEX(TENANTS!$B$4:$B${MAX_ROWS},MATCH(A{row},TENANTS!$F$4:$F${MAX_ROWS},0)),"Vacant")'
    rm[f"D{row}"] = f'=IFERROR(INDEX(TENANTS!$L$4:$L${MAX_ROWS},MATCH(A{row},TENANTS!$F$4:$F${MAX_ROWS},0)),"")'
    rm[f"E{row}"] = f'=IF(C{row}="Vacant","Vacant","Occupied")'

rm.conditional_formatting.add("E2:E13", FormulaRule(formula=['=E2="Occupied"'], fill=fill(C["light_green"]), font=font(bold=True, color=C["green"])))
rm.conditional_formatting.add("E2:E13", FormulaRule(formula=['=E2="Vacant"'], fill=fill(C["light_red"]), font=font(bold=True, color=C["red"])))

# =================================================================
# AGREEMENTS
# =================================================================
ag = wb.create_sheet("AGREEMENTS")
ag.sheet_view.showGridLines = False

ag.merge_cells("A1:Q1")
c = ag["A1"]
c.value = "LEASE AGREEMENTS & RENEWALS"
c.fill = fill(C["teal"])
c.font = font(bold=True, color=C["white"], size=14)
c.alignment = align("center")
ag.row_dimensions[1].height = 32

ag.merge_cells("A2:Q2")
c = ag["A2"]
c.value = "Each row = one agreement period. A renewed lease gets a new row. Link Tenant ID to the TENANTS sheet."
c.fill = fill(C["dark_grey"])
c.font = font(color=C["white"], size=9, italic=True)
c.alignment = align("center")
ag.row_dimensions[2].height = 18

AG_HEADERS = [
    ("A", "Agr. ID", 10), ("B", "Tenant ID", 11), ("C", "Tenant Name", 22),
    ("D", "Unit No.", 10), ("E", "Original Start", 13), ("F", "Lease Start", 13),
    ("G", "Lease End", 13), ("H", "Renewal?", 10), ("I", "Renewal No.", 11),
    ("J", "Previous Agr. Ref", 16), ("K", "Term (months)", 13), ("L", "Annual Rent", 14),
    ("M", "Monthly Rent", 13), ("N", "Currency", 10), ("O", "Rent Increment%", 13),
    ("P", "Next Review Date", 15), ("Q", "Notes", 28),
]
for col_letter, header, width in AG_HEADERS:
    c = ag[f"{col_letter}3"]
    c.value = header
    c.fill = fill(C["teal"])
    c.font = font(bold=True, color=C["white"], size=9)
    c.alignment = align("center")
    c.border = thick_bottom()
    ag.column_dimensions[col_letter].width = width
ag.row_dimensions[3].height = 22
ag.freeze_panes = "A4"

AG_EXAMPLE = {"A": "AGR-001", "B": "T-001", "C": "Kwame Mensah Enterprises", "D": "A-101",
              "H": "No", "I": 0, "J": "", "L": 42000, "N": "GHS", "O": 0.05}

for row in range(4, 54):
    is_example = (row == 4)
    is_even = (row % 2 == 0)
    for col_letter, _, _ in AG_HEADERS:
        c = ag[f"{col_letter}{row}"]
        c.font = font(size=9)
        c.border = border()
        c.alignment = align()
        if col_letter not in ["K", "M", "P"]:
            if is_example and col_letter in AG_EXAMPLE:
                c.value = AG_EXAMPLE[col_letter]
            c.fill = fill(C["yellow_inp"]) if is_example else fill(C["off_white"] if is_even else C["white"])
            if is_example:
                c.font = font(color="00008B", size=9)

    k = ag[f"K{row}"]
    k.value = f'=IFERROR(DATEDIF(F{row},G{row},"M"),"")'
    k.font = font(color="1A5276", size=9)
    k.alignment = align("center")
    k.fill = fill(C["light_blue"])

    m = ag[f"M{row}"]
    m.value = f'=IFERROR(IF(L{row}="","",L{row}/12),"")'
    m.font = font(color="1A5276", size=9)
    m.alignment = align("center")
    m.number_format = "#,##0.00"
    m.fill = fill(C["light_blue"])

    p = ag[f"P{row}"]
    p.value = f'=IFERROR(IF(G{row}="","",G{row}-365),"")'
    p.font = font(color="1A5276", size=9)
    p.alignment = align("center")
    p.number_format = "DD-MMM-YYYY"
    p.fill = fill(C["light_blue"])

for row in range(4, 54):
    for col in ["E", "F", "G"]:
        ag[f"{col}{row}"].number_format = "DD-MMM-YYYY"
    ag[f"L{row}"].number_format = "#,##0.00"
    ag[f"O{row}"].number_format = "0.0%"

ag["E4"].value = datetime.date(2024, 1, 1)
ag["F4"].value = datetime.date(2024, 1, 1)
ag["G4"].value = datetime.date(2024, 12, 31)
ag["O4"].value = 0.05

ag.conditional_formatting.add("H4:H53", FormulaRule(formula=['=H4="Yes"'], fill=fill(C["light_teal"]), font=font(color=C["teal"], bold=True)))
add_validation(ag, "H", 4, 53, lookup_range("YesNo"))
add_validation(ag, "N", 4, 53, lookup_range("Currency"))

# =================================================================
# DASHBOARD
# =================================================================
db = wb.create_sheet("DASHBOARD", 0)
db.sheet_view.showGridLines = False

db.merge_cells("A1:N1")
c = db["A1"]
c.value = "TENANT MANAGEMENT DASHBOARD"
c.fill = fill(C["navy"])
c.font = font(bold=True, color=C["white"], size=16)
c.alignment = align("center")
db.row_dimensions[1].height = 40

db.merge_cells("A2:N2")
c = db["A2"]
c.value = '=_xlfn.TEXTJOIN("  -  ",TRUE,"Report Date:",TEXT(TODAY(),"DD MMMM YYYY"))'
c.fill = fill(C["dark_grey"])
c.font = font(color=C["white"], size=10, italic=True)
c.alignment = align("center")
db.row_dimensions[2].height = 20

db.merge_cells("A4:N4")
c = db["A4"]
c.value = "  KEY PERFORMANCE INDICATORS"
c.fill = fill(C["teal"])
c.font = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[4].height = 24

KPI_DEFS = [
    ("A", "B", "TOTAL UNITS", f'=COUNTA(TENANTS!B4:B{MAX_ROWS})', "Registered tenants", C["navy"], C["white"]),
    ("C", "D", "ACTIVE LEASES", f'=COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active")+COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active - Review")', "Currently running", C["teal"], C["white"]),
    ("E", "F", "EXPIRING <= 30 DAYS", f'=COUNTIF(TENANTS!R4:R{MAX_ROWS},"Expiring Soon")', "Urgent renewals needed", C["amber"], C["white"]),
    ("G", "H", "EXPIRED", f'=COUNTIF(TENANTS!R4:R{MAX_ROWS},"Expired")', "Requires action", C["red"], C["white"]),
    ("I", "J", "TOTAL MONTHLY RENT", f'=SUMIF(TENANTS!R4:R{MAX_ROWS},"Active",TENANTS!L4:L{MAX_ROWS})+SUMIF(TENANTS!R4:R{MAX_ROWS},"Active - Review",TENANTS!L4:L{MAX_ROWS})', "Active leases (GHS)", C["green"], C["white"]),
    ("K", "L", "TOTAL ANNUAL RENT", f'=(SUMIF(TENANTS!R4:R{MAX_ROWS},"Active",TENANTS!L4:L{MAX_ROWS})+SUMIF(TENANTS!R4:R{MAX_ROWS},"Active - Review",TENANTS!L4:L{MAX_ROWS}))*12', "Projected full year (GHS)", "0D6E4E", C["white"]),
    ("M", "N", "OCCUPANCY RATE", f'=IFERROR((COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active")+COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active - Review"))/COUNTA(TENANTS!B4:B{MAX_ROWS}),0)', "Active / Total registered", C["navy"], C["white"]),
]

for cs, ce, label, formula, note, bg, fg in KPI_DEFS:
    db.merge_cells(f"{cs}5:{ce}5")
    c = db[f"{cs}5"]
    c.value = label
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=8)
    c.alignment = align("center")

    db.merge_cells(f"{cs}6:{ce}8")
    c = db[f"{cs}6"]
    c.value = formula
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=20)
    c.alignment = align("center")
    if "RATE" in label:
        c.number_format = "0.0%"
    elif "RENT" in label:
        c.number_format = "#,##0"

    db.merge_cells(f"{cs}9:{ce}9")
    c = db[f"{cs}9"]
    c.value = note
    c.fill = fill(bg)
    c.font = font(color=fg, size=8, italic=True)
    c.alignment = align("center")

for row in range(5, 10):
    db.row_dimensions[row].height = 20 if row != 7 else 28
for col_letter in "ABCDEFGHIJKLMN":
    db.column_dimensions[col_letter].width = 14

db.merge_cells("A11:N11")
c = db["A11"]
c.value = "  EXPIRY WATCH  -  Leases expiring within 90 days"
c.fill = fill(C["amber"])
c.font = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[11].height = 22

EW_COLS = [("A", "B", "Tenant ID"), ("C", "E", "Tenant Name"), ("F", "G", "Unit No."),
           ("H", "I", "Lease End"), ("J", "K", "Days Left"), ("L", "N", "Status")]
for cs, ce, label in EW_COLS:
    db.merge_cells(f"{cs}12:{ce}12")
    c = db[f"{cs}12"]
    c.value = label
    c.fill = fill(C["amber"])
    c.font = font(bold=True, color=C["white"], size=9)
    c.alignment = align("center")
    c.border = thick_bottom()
db.row_dimensions[12].height = 20

for i, row in enumerate(range(13, 23), 1):
    db.merge_cells(f"A{row}:B{row}")
    c = db[f"A{row}"]
    c.value = (f'=IFERROR(INDEX(TENANTS!A$4:A${MAX_ROWS},MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"-")')
    c.font = font(size=9); c.border = border(); c.alignment = align("center")
    c.fill = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.merge_cells(f"C{row}:E{row}")
    c = db[f"C{row}"]
    c.value = (f'=IFERROR(INDEX(TENANTS!B$4:B${MAX_ROWS},MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"-")')
    c.font = font(size=9, bold=True); c.border = border()
    c.fill = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.merge_cells(f"F{row}:G{row}")
    c = db[f"F{row}"]
    c.value = (f'=IFERROR(INDEX(TENANTS!F$4:F${MAX_ROWS},MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"-")')
    c.font = font(size=9); c.border = border(); c.alignment = align("center")
    c.fill = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.merge_cells(f"H{row}:I{row}")
    c = db[f"H{row}"]
    c.value = (f'=IFERROR(INDEX(TENANTS!J$4:J${MAX_ROWS},MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"-")')
    c.font = font(size=9); c.border = border(); c.alignment = align("center")
    c.number_format = "DD-MMM-YYYY"
    c.fill = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.merge_cells(f"J{row}:K{row}")
    c = db[f"J{row}"]
    c.value = (f'=IFERROR(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),TENANTS!Q$4:Q${MAX_ROWS}),{i}),"-")')
    c.font = font(size=9, bold=True, color=C["amber"]); c.border = border(); c.alignment = align("center")
    c.fill = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.merge_cells(f"L{row}:N{row}")
    c = db[f"L{row}"]
    c.value = (f'=IFERROR(INDEX(TENANTS!R$4:R${MAX_ROWS},MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"-")')
    c.font = font(size=9, bold=True); c.border = border(); c.alignment = align("center")
    c.fill = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.row_dimensions[row].height = 18

db.merge_cells("A24:N24")
c = db["A24"]
c.value = "  PORTFOLIO OVERVIEW  -  Lease status breakdown"
c.fill = fill(C["navy"])
c.font = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[24].height = 22

STATUS_BREAKDOWN = [
    ("Active", C["green"], C["white"]), ("Active - Review", "0D6E4E", C["white"]),
    ("Expiring Soon", C["amber"], C["white"]), ("Expired", C["red"], C["white"]),
    ("Terminated", C["dark_grey"], C["white"]), ("Vacant", C["mid_grey"], "000000"),
]
for j, col_letter in enumerate(["A", "C", "E", "G", "I", "K"]):
    cs = col_letter
    ce = ["B", "D", "F", "H", "J", "L"][j]
    status, bg, fg = STATUS_BREAKDOWN[j]
    db.merge_cells(f"{cs}25:{ce}25")
    c = db[f"{cs}25"]
    c.value = status; c.fill = fill(bg); c.font = font(bold=True, color=fg, size=9)
    c.alignment = align("center"); c.border = border()

    db.merge_cells(f"{cs}26:{ce}27")
    c = db[f"{cs}26"]
    c.value = f'=COUNTIF(TENANTS!R$4:R${MAX_ROWS},"{status}")'
    c.fill = fill(bg); c.font = font(bold=True, color=fg, size=18)
    c.alignment = align("center"); c.border = border()

for row in [25, 26, 27]:
    db.row_dimensions[row].height = 22

db.merge_cells("A29:N29")
c = db["A29"]
c.value = "  HOW TO USE THIS WORKBOOK"
c.fill = fill(C["dark_grey"])
c.font = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[29].height = 22

INSTRUCTIONS = [
    ("1", "Use the TENANT FORM sheet to enter a new tenant's details in the yellow cells, then copy the row into TENANTS (or use Add once macros are enabled)."),
    ("2", "Go to the AGREEMENTS sheet to record each lease agreement. Add a new row for each renewal - do NOT overwrite old rows."),
    ("3", "The DASHBOARD updates automatically. The Expiry Watch section shows the soonest-expiring leases within 90 days."),
    ("4", "Status is automatic: Active -> Expiring Soon (<=30 days) -> Expired. 'Active - Review' means the lease ends within 3 months."),
    ("5", "For a renewal: add a new row in AGREEMENTS with 'Renewal?' = Yes, Renewal No. = 1 (or 2 for second renewal). Update the Lease End date in TENANTS."),
    ("6", "Yellow cells = you fill in - Blue-text cells = formula, do not edit - Drop-down arrows appear when you click a cell in the relevant columns."),
]
for k, (num, text) in enumerate(INSTRUCTIONS):
    row = 30 + k
    db[f"A{row}"].value = num
    db[f"A{row}"].font = font(bold=True, color=C["white"], size=10)
    db[f"A{row}"].fill = fill(C["teal"])
    db[f"A{row}"].alignment = align("center")
    db[f"A{row}"].border = border()

    db.merge_cells(f"B{row}:N{row}")
    c = db[f"B{row}"]
    c.value = text
    c.font = font(size=9)
    c.fill = fill(C["off_white"] if k % 2 == 0 else C["white"])
    c.alignment = align("left", wrap=True)
    c.border = border()
    db.row_dimensions[row].height = 28

for ws in [db, tn, ag, form]:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

wb.save(OUT)
print(f"Saved: {OUT}")