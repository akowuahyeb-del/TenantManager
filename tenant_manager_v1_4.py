"""
Tenant Management Excel Workbook
=================================
Sheets:
  1. DASHBOARD   – live KPIs, expiry alerts, occupancy
  2. TENANTS     – master register with all tenant details
  3. AGREEMENTS  – lease terms, renewals, payment tracking
  4. LOOKUP      – drop-down lists (hidden helper sheet)
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import datetime

OUT = "Tenant_Manager.xlsx"

# ── PALETTE ──────────────────────────────────────────────────────
C = {
    "navy":       "1F3864",
    "teal":       "17A589",
    "amber":      "E67E22",
    "red":        "E74C3C",
    "green":      "1E8449",
    "light_teal": "D1F2EB",
    "light_amber":"FDEBD0",
    "light_red":  "FADBD8",
    "light_green":"D5F5E3",
    "light_blue": "D6EAF8",
    "white":      "FFFFFF",
    "off_white":  "F8F9FA",
    "mid_grey":   "BFC9CA",
    "dark_grey":  "566573",
    "yellow_inp": "FFF9C4",
    "label_bg": "EAF2FF",   # user-input cells
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def border(style="thin"):
    s = Side(border_style=style, color="BFC9CA")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thin = Side(border_style="thin",   color="BFC9CA")
    thk  = Side(border_style="medium", color="1F3864")
    return Border(left=thin, right=thin, top=thin, bottom=thk)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header_row(ws, row, cols, bg=C["navy"], fg=C["white"], sz=10):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill  = fill(bg)
        c.font  = font(bold=True, color=fg, size=sz)
        c.alignment = align("center")
        c.border = border()

def style_input(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.fill      = fill(C["yellow_inp"])
    c.font      = font(color="00008B")   # blue = user edits here
    c.border    = border()
    c.alignment = align()

def alternating(ws, row, col_start, col_end, even):
    bg = C["off_white"] if even else C["white"]
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = fill(bg)
        c.border = border()
        c.alignment = align()

# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# SHEET 4: LOOKUP  (helper – hidden)
# ═══════════════════════════════════════════════════════════════
lk = wb.active
lk.title = "LOOKUP"

UNIT_TYPES  = ["Apartment","Office","Shop","Warehouse","Studio","Penthouse","Other"]
STATUS_LIST = ["Active","Pending","Expired","Terminated","Vacant"]
FREQ_LIST   = ["Monthly","Quarterly","Semi-Annual","Annual"]
CURRENCY    = ["GHS","USD","GBP","EUR","NGN"]
PAYMENT_M   = ["Cash","Bank Transfer","Mobile Money","Cheque","Standing Order"]
YES_NO      = ["Yes","No"]

lists = {
    "UnitTypes":  UNIT_TYPES,
    "StatusList": STATUS_LIST,
    "FreqList":   FREQ_LIST,
    "Currency":   CURRENCY,
    "PayMethod":  PAYMENT_M,
    "YesNo":      YES_NO,
}

lk["A1"].value = "List Name"
lk["B1"].value = "Values"
r = 2
named_ranges = {}
for name, items in lists.items():
    start_row = r
    for item in items:
        lk.cell(row=r, column=2).value = item
        r += 1
    end_row = r - 1
    named_ranges[name] = f"LOOKUP!$B${start_row}:$B${end_row}"
    lk.cell(row=start_row, column=1).value = name

lk.sheet_state = "hidden"


# ═══════════════════════════════════════════════════════════════
# SHEET: TENANT FORM (UPGRADED)
# ═══════════════════════════════════════════════════════════════

form = wb.create_sheet("TENANT FORM")
form.sheet_view.showGridLines = False

# Column widths
form.column_dimensions["A"].width = 3
form.column_dimensions["B"].width = 25
form.column_dimensions["C"].width = 3
form.column_dimensions["D"].width = 22
form.column_dimensions["E"].width = 18
form.column_dimensions["F"].width = 18
form.column_dimensions["G"].width = 18

# ─────────────────────────
# HEADER
# ─────────────────────────

form.merge_cells("A1:G1")
c = form["A1"]
c.value = "🏢 PROPERTY MANAGEMENT SYSTEM"
c.fill = fill(C["navy"])
c.font = font(bold=True, color=C["white"], size=13)
c.alignment = align("center")

form.merge_cells("A2:G2")
c = form["A2"]
c.value = "TENANT REGISTRATION FORM"
c.fill = fill(C["teal"])
c.font = font(bold=True, color=C["white"], size=16)
c.alignment = align("center")

form.merge_cells("A3:G3")
c = form["A3"]
c.value = "Complete all fields • Yellow cells are editable"
c.fill = fill(C["dark_grey"])
c.font = font(color=C["white"], size=9, italic=True)
c.alignment = align("center")

# ─────────────────────────
# HELPER
# ─────────────────────────

def form_label(row, text):
    cell = form[f"B{row}"]
    cell.value = text
    cell.fill = fill(C["label_bg"])
    cell.font = font(bold=True, color=C["navy"], size=9)
    cell.border = border()
    cell.alignment = align("right")

def input_cell(cell_ref):
    c = form[cell_ref]
    c.fill = fill(C["yellow_inp"])
    c.border = border(style="medium")
    c.font = font(size=9)

# ─────────────────────────
# TENANT DETAILS
# ─────────────────────────

fields = [
    (6,  "Tenant ID",            "D6"),
    (7,  "Full Name / Company",  "D7"),
    (8,  "Phone",                "D8"),
    (9,  "Email",                "D9"),
    (10, "ID / Reg No.",         "D10"),
    (11, "Unit Number",          "D11"),
    (12, "Unit Type",            "D12"),
    (13, "Floor / Block",        "D13"),
    (14, "Contract Start",       "D14"),
    (15, "Contract End",         "D15"),
    (16, "Monthly Rent",         "D16"),
    (17, "Currency",             "D17"),
    (18, "Payment Frequency",    "D18"),
    (19, "Deposit Paid",         "D19"),
    (20, "Payment Method",       "D20"),
]

for row, label, cell_addr in fields:
    form_label(row, label)

    form.merge_cells(
        f"{cell_addr}:{chr(ord(cell_addr[0]) + 2)}{row}"
    )

    input_cell(cell_addr)

# ✅ AUTO-GENERATED TENANT ID
form["D6"] = "T-001"

# ─────────────────────────
# NOTES SECTION
# ─────────────────────────

form.merge_cells("B23:F23")
cell = form["B23"]
cell.value = "Additional Notes / Special Conditions"
cell.fill = fill(C["navy"])
cell.font = font(bold=True, color=C["white"])
cell.alignment = align("center")

form.merge_cells("B24:F28")
notes = form["B24"]
notes.fill = fill(C["yellow_inp"])
notes.border = border(style="medium")

# ─────────────────────────
# SIGNATURES
# ─────────────────────────

form.merge_cells("B31:D31")
form["B31"] = "Tenant Signature"

form.merge_cells("E31:F31")
form["E31"] = "Property Manager"

for cell in ["B32", "E32"]:
    form[cell].fill = fill(C["yellow_inp"])
    form[cell].border = border(style="medium")

# Footer
form.merge_cells("A35:G35")
c = form["A35"]
c.value = "Transfer approved tenant records into TENANTS sheet"
c.fill = fill(C["dark_grey"])
c.font = font(color=C["white"], size=8, italic=True)
c.alignment = align("center")


# ─────────────────────────
# TENANT FORM DROPDOWNS
# ─────────────────────────

# Unit Type
unit_type_dv = DataValidation(
    type="list",
    formula1='"Office,Shop,Apartment,Warehouse,Studio"'
)
form.add_data_validation(unit_type_dv)
unit_type_dv.add("D12")

# Currency
currency_dv = DataValidation(
    type="list",
    formula1='"GHS,USD,EUR,GBP"'
)
form.add_data_validation(currency_dv)
currency_dv.add("D17")

# Payment Frequency
payment_freq_dv = DataValidation(
    type="list",
    formula1='"Monthly,Quarterly,Semi-Annual,Annual"'
)
form.add_data_validation(payment_freq_dv)
payment_freq_dv.add("D18")

# Payment Method
payment_method_dv = DataValidation(
    type="list",
    formula1='"Cash,Bank Transfer,Mobile Money,Cheque"'
)
form.add_data_validation(payment_method_dv)
payment_method_dv.add("D20")


# ═══════════════════════════════════════════════════════════════
# SHEET 2: TENANTS
# ═══════════════════════════════════════════════════════════════
tn = wb.create_sheet("TENANTS")
tn.sheet_view.showGridLines = False

# ── Title bar ──
tn.merge_cells("A1:R1")
c = tn["A1"]
c.value     = "🏢  TENANT REGISTER"
c.fill      = fill(C["navy"])
c.font      = font(bold=True, color=C["white"], size=14)
c.alignment = align("center")
tn.row_dimensions[1].height = 32

tn.merge_cells("A2:R2")
c = tn["A2"]
c.value     = "Yellow cells = data entry  ·  Blue text = formula (do not edit)  ·  Status updates automatically"
c.fill      = fill(C["dark_grey"])
c.font      = font(color=C["white"], size=9, italic=True)
c.alignment = align("center")
tn.row_dimensions[2].height = 18

# ── Column headers (row 3) ──
T_HEADERS = [
    ("A", "Tenant ID",         11), ("B", "Full Name / Company",  22),
    ("C", "Phone",             14), ("D", "Email",                24),
    ("E", "ID / Reg No.",      14), ("F", "Unit No.",             10),
    ("G", "Unit Type",         12), ("H", "Floor / Block",        12),
    ("I", "Contract Start",    13), ("J", "Contract End",         13),
    ("K", "Term (months)",     13), ("L", "Monthly Rent",         13),
    ("M", "Currency",          10), ("N", "Payment Freq.",        13),
    ("O", "Deposit Paid",      13), ("P", "Payment Method",       14),
    ("Q", "Days to Expiry",    14), ("R", "Status",               12),
]

for col_letter, header, width in T_HEADERS:
    c = tn[f"{col_letter}3"]
    c.value     = header
    c.fill      = fill(C["navy"])
    c.font      = font(bold=True, color=C["white"], size=9)
    c.alignment = align("center")
    c.border    = thick_bottom()
    tn.column_dimensions[col_letter].width = width

tn.row_dimensions[3].height = 22
tn.freeze_panes = "A4"

# ── Example row 4 then blank rows 5-53 ──
EXAMPLE = {
    "A": "T-001",
    "B": "Kwame Mensah Enterprises",
    "C": "0244123456",
    "D": "kwame@example.com",
    "E": "GH-0012345",
    "F": "A-101",
    "G": "Office",
    "H": "1st / Block A",
    "I": "2024-01-01",
    "J": "2024-12-31",
    "L": 3500,
    "M": "GHS",
    "N": "Monthly",
    "O": 7000,
    "P": "Bank Transfer",
}

input_cols  = ["A","B","C","D","E","F","G","H","I","J","L","M","N","O","P"]
formula_cols = ["K","Q","R"]
MAX_ROWS = 53

for row in range(4, MAX_ROWS + 1):
    is_example = (row == 4)
    is_even    = (row % 2 == 0)

    for col_letter, _, _ in T_HEADERS:
        c       = tn[f"{col_letter}{row}"]
        c.font  = font(size=9)
        c.border = border()
        c.alignment = align()

        if col_letter in input_cols:
            if is_example and col_letter in EXAMPLE:
                c.value = EXAMPLE[col_letter]
            c.fill = fill(C["yellow_inp"]) if is_example else fill(C["off_white"] if is_even else C["white"])
            c.font = font(color="00008B" if is_example else "000000", size=9)

    # Formula: Term in months (K)
    k = tn[f"K{row}"]
    k.value     = f'=IFERROR(DATEDIF(I{row},J{row},"M"),"")'
    k.font      = font(color="1A5276", size=9)   # blue = formula
    k.alignment = align("center")
    k.fill      = fill(C["light_blue"])

    # Formula: Days to Expiry (Q)
    q = tn[f"Q{row}"]
    q.value     = f'=IFERROR(IF(J{row}="","",J{row}-TODAY()),"")'
    q.font      = font(color="1A5276", size=9)
    q.alignment = align("center")
    q.number_format = "0"
    q.fill      = fill(C["light_blue"])

    # Formula: Status (R) — auto from days to expiry
    r_cell = tn[f"R{row}"]
    r_cell.value = (
        f'=IFERROR(IF(B{row}="",'
        f'"Vacant",'
        f'IF(Q{row}<0,"Expired",'
        f'IF(Q{row}<=30,"Expiring Soon",'
        f'IF(Q{row}<=90,"Active - Review","Active")))),"Vacant")'
    )
    r_cell.font      = font(bold=True, size=9)
    r_cell.alignment = align("center")
    r_cell.fill      = fill(C["light_blue"])

# ── Date format for I and J ──
for row in range(4, MAX_ROWS + 1):
    tn[f"I{row}"].number_format = "DD-MMM-YYYY"
    tn[f"J{row}"].number_format = "DD-MMM-YYYY"
    tn[f"L{row}"].number_format = '#,##0.00'
    tn[f"O{row}"].number_format = '#,##0.00'

# ── Conditional formatting on Status (R) ──
red_rule = FormulaRule(
    formula=[f'=OR(R4="Expired",R4="Terminated")'],
    fill=fill(C["light_red"]),
    font=font(bold=True, color=C["red"])
)
amber_rule = FormulaRule(
    formula=[f'=R4="Expiring Soon"'],
    fill=fill(C["light_amber"]),
    font=font(bold=True, color=C["amber"])
)
green_rule = FormulaRule(
    formula=[f'=OR(R4="Active",R4="Active - Review")'],
    fill=fill(C["light_green"]),
    font=font(bold=True, color=C["green"])
)
tn.conditional_formatting.add(f"R4:R{MAX_ROWS}", red_rule)
tn.conditional_formatting.add(f"R4:R{MAX_ROWS}", amber_rule)
tn.conditional_formatting.add(f"R4:R{MAX_ROWS}", green_rule)

# Conditional on Days to Expiry (Q) — colour scale
tn.conditional_formatting.add(
    f"Q4:Q{MAX_ROWS}",
    ColorScaleRule(
        start_type="num", start_value=0,   start_color="E74C3C",
        mid_type="num",   mid_value=90,    mid_color="F39C12",
        end_type="num",   end_value=365,   end_color="27AE60",
    )
)

# ── Drop-down validations ──
def add_validation(ws, col_letter, row_start, row_end, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)

add_validation(tn, "G", 4, MAX_ROWS, f'LOOKUP!$B$2:$B${1+len(UNIT_TYPES)}')
add_validation(tn, "M", 4, MAX_ROWS, f'LOOKUP!$B${2+len(UNIT_TYPES)+len(STATUS_LIST)}:$B${1+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)+len(CURRENCY)}')
add_validation(tn, "N", 4, MAX_ROWS, f'LOOKUP!$B${2+len(UNIT_TYPES)+len(STATUS_LIST)}:$B${1+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)}')
add_validation(tn, "P", 4, MAX_ROWS, f'LOOKUP!$B${2+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)+len(CURRENCY)}:$B${1+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)+len(CURRENCY)+len(PAYMENT_M)}')


# ═══════════════════════════════════════════════════════════════
# SHEET: ROOMS
# ═══════════════════════════════════════════════════════════════

rm = wb.create_sheet("ROOMS")

ROOM_HEADERS = [
    ("A", "Room No", 12),
    ("B", "Tenant ID", 12),
    ("C", "Tenant Name", 25),
    ("D", "Monthly Rent", 15),
    ("E", "Status", 15),
]

# Headers
for col, header, width in ROOM_HEADERS:
    rm[f"{col}1"] = header
    rm.column_dimensions[col].width = width

# 12 rooms
for i in range(1, 13):
    rm[f"A{i+1}"] = f"RM-{i:02d}"




# ═══════════════════════════════════════════════════════════════
# SHEET 3: AGREEMENTS  (renewals + payment schedule)
# ═══════════════════════════════════════════════════════════════
ag = wb.create_sheet("AGREEMENTS")
ag.sheet_view.showGridLines = False

ag.merge_cells("A1:P1")
c = ag["A1"]
c.value     = "📋  LEASE AGREEMENTS & RENEWALS"
c.fill      = fill(C["teal"])
c.font      = font(bold=True, color=C["white"], size=14)
c.alignment = align("center")
ag.row_dimensions[1].height = 32

ag.merge_cells("A2:P2")
c = ag["A2"]
c.value     = "Each row = one agreement period. A renewed lease gets a new row. Link Tenant ID to the TENANTS sheet."
c.fill      = fill(C["dark_grey"])
c.font      = font(color=C["white"], size=9, italic=True)
c.alignment = align("center")
ag.row_dimensions[2].height = 18

AG_HEADERS = [
    ("A", "Agr. ID",          10), ("B", "Tenant ID",       11),
    ("C", "Tenant Name",      22), ("D", "Unit No.",        10),
    ("E", "Original Start",   13), ("F", "Lease Start",     13),
    ("G", "Lease End",        13), ("H", "Renewal?",        10),
    ("I", "Renewal No.",      11), ("J", "Term (months)",   13),
    ("K", "Annual Rent",      14), ("L", "Monthly Rent",    13),
    ("M", "Currency",         10), ("N", "Rent Increment%", 13),
    ("O", "Next Review Date", 15), ("P", "Notes",           28),
]

for col_letter, header, width in AG_HEADERS:
    c = ag[f"{col_letter}3"]
    c.value     = header
    c.fill      = fill(C["teal"])
    c.font      = font(bold=True, color=C["white"], size=9)
    c.alignment = align("center")
    c.border    = thick_bottom()
    ag.column_dimensions[col_letter].width = width

ag.row_dimensions[3].height = 22
ag.freeze_panes = "A4"

# Example agreement row
AG_EXAMPLE = {
    "A": "AGR-001", "B": "T-001", "C": "Kwame Mensah Enterprises",
    "D": "A-101",   "H": "No",    "I": 0,
    "K": 42000,     "M": "GHS",   "N": 0.05,
}

for row in range(4, 54):
    is_example = (row == 4)
    is_even    = (row % 2 == 0)

    for col_letter, _, _ in AG_HEADERS:
        c       = ag[f"{col_letter}{row}"]
        c.font  = font(size=9)
        c.border = border()
        c.alignment = align()

        if col_letter not in ["J","L","O"]:
            if is_example and col_letter in AG_EXAMPLE:
                c.value = AG_EXAMPLE[col_letter]
            if col_letter not in ["J","L","O"]:
                c.fill = fill(C["yellow_inp"]) if is_example else fill(C["off_white"] if is_even else C["white"])
                if is_example:
                    c.font = font(color="00008B", size=9)

    # J = Term months formula
    j = ag[f"J{row}"]
    j.value     = f'=IFERROR(DATEDIF(F{row},G{row},"M"),"")'
    j.font      = font(color="1A5276", size=9)
    j.alignment = align("center")
    j.fill      = fill(C["light_blue"])

    # L = Monthly rent from annual
    l = ag[f"L{row}"]
    l.value     = f'=IFERROR(IF(K{row}="","",K{row}/12),"")'
    l.font      = font(color="1A5276", size=9)
    l.alignment = align("center")
    l.number_format = '#,##0.00'
    l.fill      = fill(C["light_blue"])

    # O = Next review date (1 year before end)
    o = ag[f"O{row}"]
    o.value     = f'=IFERROR(IF(G{row}="","",G{row}-365),"")'
    o.font      = font(color="1A5276", size=9)
    o.alignment = align("center")
    o.number_format = "DD-MMM-YYYY"
    o.fill      = fill(C["light_blue"])

for row in range(4, 54):
    for col in ["E","F","G"]:
        ag[f"{col}{row}"].number_format = "DD-MMM-YYYY"
    ag[f"K{row}"].number_format = '#,##0.00'
    ag[f"N{row}"].number_format = '0.0%'

# Set example dates
if True:
    ag["E4"].value = datetime.date(2024, 1, 1)
    ag["F4"].value = datetime.date(2024, 1, 1)
    ag["G4"].value = datetime.date(2024, 12, 31)
    ag["N4"].value = 0.05

# Conditional: highlight renewals
renewal_rule = FormulaRule(
    formula=['=H4="Yes"'],
    fill=fill(C["light_teal"]),
    font=font(color=C["teal"], bold=True)
)
ag.conditional_formatting.add("H4:H53", renewal_rule)

# Drop-downs on agreements
add_validation(ag, "H", 4, 53,
    f'LOOKUP!$B${2+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)+len(CURRENCY)+len(PAYMENT_M)}:'
    f'$B${1+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)+len(CURRENCY)+len(PAYMENT_M)+len(YES_NO)}'
)
add_validation(ag, "M", 4, 53,
    f'LOOKUP!$B${2+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)}:'
    f'$B${1+len(UNIT_TYPES)+len(STATUS_LIST)+len(FREQ_LIST)+len(CURRENCY)}'
)

# ═══════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════
db = wb.create_sheet("DASHBOARD", 0)
db.sheet_view.showGridLines = False

# Big title
db.merge_cells("A1:N1")
c = db["A1"]
c.value     = "🏢   TENANT MANAGEMENT DASHBOARD"
c.fill      = fill(C["navy"])
c.font      = font(bold=True, color=C["white"], size=16)
c.alignment = align("center")
db.row_dimensions[1].height = 40

db.merge_cells("A2:N2")
c = db["A2"]
c.value     = '=_xlfn.TEXTJOIN("  ·  ",TRUE,"Report Date:",TEXT(TODAY(),"DD MMMM YYYY"))'
c.fill      = fill(C["dark_grey"])
c.font      = font(color=C["white"], size=10, italic=True)
c.alignment = align("center")
db.row_dimensions[2].height = 20

# ── Row 4: Section label ──
db.merge_cells("A4:N4")
c = db["A4"]
c.value     = "  KEY PERFORMANCE INDICATORS"
c.fill      = fill(C["teal"])
c.font      = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[4].height = 24

# ── KPI Cards  (rows 5-9) ──
KPI_DEFS = [
    # (col_start, col_end, label, formula, note, bg, text_color)
    ("A", "B", "TOTAL UNITS",
     f'=COUNTA(TENANTS!B4:B{MAX_ROWS})',
     "Registered tenants", C["navy"], C["white"]),

    ("C", "D", "ACTIVE LEASES",
     f'=COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active")+COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active - Review")',
     "Currently running", C["teal"], C["white"]),

    ("E", "F", "EXPIRING ≤ 30 DAYS",
     f'=COUNTIF(TENANTS!R4:R{MAX_ROWS},"Expiring Soon")',
     "Urgent renewals needed", C["amber"], C["white"]),

    ("G", "H", "EXPIRED",
     f'=COUNTIF(TENANTS!R4:R{MAX_ROWS},"Expired")',
     "Requires action", C["red"], C["white"]),

    ("I", "J", "TOTAL MONTHLY RENT",
     f'=SUMIF(TENANTS!R4:R{MAX_ROWS},"Active",TENANTS!L4:L{MAX_ROWS})+SUMIF(TENANTS!R4:R{MAX_ROWS},"Active - Review",TENANTS!L4:L{MAX_ROWS})',
     "Active leases (GHS)", C["green"], C["white"]),

    ("K", "L", "TOTAL ANNUAL RENT",
     f'=(SUMIF(TENANTS!R4:R{MAX_ROWS},"Active",TENANTS!L4:L{MAX_ROWS})+SUMIF(TENANTS!R4:R{MAX_ROWS},"Active - Review",TENANTS!L4:L{MAX_ROWS}))*12',
     "Projected full year (GHS)", "0D6E4E", C["white"]),

    ("M", "N", "OCCUPANCY RATE",
     f'=IFERROR((COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active")+COUNTIF(TENANTS!R4:R{MAX_ROWS},"Active - Review"))/COUNTA(TENANTS!B4:B{MAX_ROWS}),0)',
     "Active / Total registered", C["navy"], C["white"]),
]

for cs, ce, label, formula, note, bg, fg in KPI_DEFS:
    # Label row (5)
    db.merge_cells(f"{cs}5:{ce}5")
    c = db[f"{cs}5"]
    c.value     = label
    c.fill      = fill(bg)
    c.font      = font(bold=True, color=fg, size=8)
    c.alignment = align("center")

    # Value row (6-8)
    db.merge_cells(f"{cs}6:{ce}8")
    c = db[f"{cs}6"]
    c.value     = formula
    c.fill      = fill(bg)
    c.font      = font(bold=True, color=fg, size=20)
    c.alignment = align("center")
    if "RATE" in label:
        c.number_format = "0.0%"
    elif "RENT" in label:
        c.number_format = "#,##0"

    # Note row (9)
    db.merge_cells(f"{cs}9:{ce}9")
    c = db[f"{cs}9"]
    c.value     = note
    c.fill      = fill(bg)
    c.font      = font(color=fg, size=8, italic=True)
    c.alignment = align("center")

for row in range(5, 10):
    db.row_dimensions[row].height = 20 if row != 7 else 28

# Set uniform column widths for KPI cards
for col_letter in "ABCDEFGHIJKLMN":
    db.column_dimensions[col_letter].width = 14

# ── Row 11: Divider ──
db.merge_cells("A11:N11")
c = db["A11"]
c.value = "  EXPIRY WATCH  —  Leases expiring within 90 days"
c.fill  = fill(C["amber"])
c.font  = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[11].height = 22

# ── Expiry watch headers (row 12) ──
EW_COLS = [
    ("A","B","Tenant ID"), ("C","E","Tenant Name"),
    ("F","G","Unit No."),  ("H","I","Lease End"),
    ("J","K","Days Left"), ("L","N","Status"),
]
for cs, ce, label in EW_COLS:
    db.merge_cells(f"{cs}12:{ce}12")
    c = db[f"{cs}12"]
    c.value     = label
    c.fill      = fill(C["amber"])
    c.font      = font(bold=True, color=C["white"], size=9)
    c.alignment = align("center")
    c.border    = thick_bottom()
db.row_dimensions[12].height = 20

# ── Expiry watch data rows (13-22) using array-like IFERROR+INDEX/MATCH ──
# We rank by days-to-expiry using a helper approach.
# Each row shows the nth-soonest-expiring tenant with 0-90 days left.
for i, row in enumerate(range(13, 23), 1):
    # The formula finds the tenant whose days-to-expiry is the i-th smallest
    # positive value ≤ 90.
    # Tenant ID
    db.merge_cells(f"A{row}:B{row}")
    c = db[f"A{row}"]
    c.value = (
        f'=IFERROR(INDEX(TENANTS!A$4:A${MAX_ROWS},'
        f'MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),'
        f'TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"—")'
    )
    c.font      = font(size=9)
    c.border    = border()
    c.alignment = align("center")
    c.fill      = fill(C["off_white"] if i % 2 == 0 else C["white"])

    # Tenant Name
    db.merge_cells(f"C{row}:E{row}")
    c = db[f"C{row}"]
    c.value = (
        f'=IFERROR(INDEX(TENANTS!B$4:B${MAX_ROWS},'
        f'MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),'
        f'TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"—")'
    )
    c.font  = font(size=9, bold=True)
    c.border = border()
    c.fill   = fill(C["off_white"] if i % 2 == 0 else C["white"])

    # Unit
    db.merge_cells(f"F{row}:G{row}")
    c = db[f"F{row}"]
    c.value = (
        f'=IFERROR(INDEX(TENANTS!F$4:F${MAX_ROWS},'
        f'MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),'
        f'TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"—")'
    )
    c.font   = font(size=9)
    c.border = border()
    c.alignment = align("center")
    c.fill   = fill(C["off_white"] if i % 2 == 0 else C["white"])

    # Lease End
    db.merge_cells(f"H{row}:I{row}")
    c = db[f"H{row}"]
    c.value = (
        f'=IFERROR(INDEX(TENANTS!J$4:J${MAX_ROWS},'
        f'MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),'
        f'TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"—")'
    )
    c.font   = font(size=9)
    c.border = border()
    c.alignment = align("center")
    c.number_format = "DD-MMM-YYYY"
    c.fill   = fill(C["off_white"] if i % 2 == 0 else C["white"])

    # Days Left
    db.merge_cells(f"J{row}:K{row}")
    c = db[f"J{row}"]
    c.value = (
        f'=IFERROR(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),'
        f'TENANTS!Q$4:Q${MAX_ROWS}),{i}),"—")'
    )
    c.font   = font(size=9, bold=True, color=C["amber"])
    c.border = border()
    c.alignment = align("center")
    c.fill   = fill(C["off_white"] if i % 2 == 0 else C["white"])

    # Status
    db.merge_cells(f"L{row}:N{row}")
    c = db[f"L{row}"]
    c.value = (
        f'=IFERROR(INDEX(TENANTS!R$4:R${MAX_ROWS},'
        f'MATCH(SMALL(IF((TENANTS!Q$4:Q${MAX_ROWS}>=0)*(TENANTS!Q$4:Q${MAX_ROWS}<=90),'
        f'TENANTS!Q$4:Q${MAX_ROWS}),{i}),TENANTS!Q$4:Q${MAX_ROWS},0)),"—")'
    )
    c.font   = font(size=9, bold=True)
    c.border = border()
    c.alignment = align("center")
    c.fill   = fill(C["off_white"] if i % 2 == 0 else C["white"])

    db.row_dimensions[row].height = 18

# ── Row 24: Status breakdown section ──
db.merge_cells("A24:N24")
c = db["A24"]
c.value = "  PORTFOLIO OVERVIEW  —  Lease status breakdown"
c.fill  = fill(C["navy"])
c.font  = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[24].height = 22

STATUS_BREAKDOWN = [
    ("Active",          C["green"],  C["white"]),
    ("Active - Review", "0D6E4E",    C["white"]),
    ("Expiring Soon",   C["amber"],  C["white"]),
    ("Expired",         C["red"],    C["white"]),
    ("Terminated",      C["dark_grey"], C["white"]),
    ("Vacant",          C["mid_grey"],  "000000"),
]

# headers row 25
for j, col_letter in enumerate(["A","C","E","G","I","K"]):
    cs = col_letter
    ce = ["B","D","F","H","J","L"][j]
    status, bg, fg = STATUS_BREAKDOWN[j]
    db.merge_cells(f"{cs}25:{ce}25")
    c = db[f"{cs}25"]
    c.value     = status
    c.fill      = fill(bg)
    c.font      = font(bold=True, color=fg, size=9)
    c.alignment = align("center")
    c.border    = border()

    db.merge_cells(f"{cs}26:{ce}27")
    c = db[f"{cs}26"]
    c.value     = f'=COUNTIF(TENANTS!R$4:R${MAX_ROWS},"{status}")'
    c.fill      = fill(bg)
    c.font      = font(bold=True, color=fg, size=18)
    c.alignment = align("center")
    c.border    = border()

for row in [25, 26, 27]:
    db.row_dimensions[row].height = 22

# ── Row 29: Instructions ──
db.merge_cells("A29:N29")
c = db["A29"]
c.value = "  HOW TO USE THIS WORKBOOK"
c.fill  = fill(C["dark_grey"])
c.font  = font(bold=True, color=C["white"], size=11)
c.alignment = align()
db.row_dimensions[29].height = 22

INSTRUCTIONS = [
    ("1", "Go to the TENANTS sheet to register each tenant. Fill in all yellow cells. Formula cells (blue text) calculate automatically."),
    ("2", "Go to the AGREEMENTS sheet to record each lease agreement. Add a new row for each renewal — do NOT overwrite old rows. This gives you a full history."),
    ("3", "The DASHBOARD updates automatically. The Expiry Watch section shows the 10 soonest-expiring leases within 90 days."),
    ("4", "Status is automatic: Active → Expiring Soon (≤30 days) → Expired. 'Active - Review' means the lease ends within 3 months."),
    ("5", "For a renewal: add a new row in AGREEMENTS with 'Renewal?' = Yes, Renewal No. = 1 (or 2 for second renewal, etc.). Update the Lease End date in TENANTS."),
    ("6", "Yellow cells = you fill in  ·  Blue-text cells = formula, do not edit  ·  Drop-down arrows appear when you click a cell in the relevant columns."),
]

for k, (num, text) in enumerate(INSTRUCTIONS):
    row = 30 + k
    db[f"A{row}"].value     = num
    db[f"A{row}"].font      = font(bold=True, color=C["white"], size=10)
    db[f"A{row}"].fill      = fill(C["teal"])
    db[f"A{row}"].alignment = align("center")
    db[f"A{row}"].border    = border()

    db.merge_cells(f"B{row}:N{row}")
    c = db[f"B{row}"]
    c.value     = text
    c.font      = font(size=9)
    c.fill      = fill(C["off_white"] if k % 2 == 0 else C["white"])
    c.alignment = align("left", wrap=True)
    c.border    = border()
    db.row_dimensions[row].height = 28

# ── Print settings ──
for ws in [db, tn, ag]:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

wb.save(OUT)
print(f"Saved: {OUT}")